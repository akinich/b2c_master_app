"""
woocommerce_zoho_export.py

Module Key: woocommerce_zoho_export

Fetch completed orders from WooCommerce between dates,
map item names using optional item_database.xlsx, then
export line-item CSV + summary Excel bundled into orders_export.zip.

Credentials expected in Streamlit secrets:
[woocommerce]
api_url = "https://your-site.com/wp-json/wc/v3"
consumer_key = "ck_xxxxx"
consumer_secret = "cs_xxxxx"
"""

import streamlit as st
import pandas as pd
import requests
from datetime import datetime
from dateutil.parser import parse
from collections import Counter
from io import BytesIO
from zipfile import ZipFile
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from auth.session import SessionManager
from config.database import ActivityLogger
import time

TOOL_NAME = "WooCommerce → Zoho Export"

# ------------------------
# Utilities
def to_float(x):
    try:
        if x is None or x == "":
            return 0.0
        return float(x)
    except Exception:
        return 0.0

def read_item_database(uploaded_file):
    """
    Read item database either from uploaded file or from disk (item_database.xlsx).
    Returns name_mapping dict (lowercase woo_name -> {'zoho', 'hsn', 'usage_unit'}) and dataframe (or None).
    """
    try:
        if uploaded_file:
            item_db_df = pd.read_excel(uploaded_file, dtype=str)
        else:
            item_db_df = pd.read_excel("item_database.xlsx", dtype=str)
        # Normalize headers to lowercase and stripped
        item_db_df.columns = [str(col).strip().lower() for col in item_db_df.columns]

        # Validate required columns (best-effort)
        required_columns = ["woocommerce name", "zoho name", "hsn", "usage unit"]
        for col in required_columns:
            if col not in item_db_df.columns:
                st.warning(f"item_database.xlsx missing expected column: '{col}' (proceeding with available data).")

        # Build mapping
        name_mapping = {}
        for _, row in item_db_df.iterrows():
            woo_raw = row.get("woocommerce name", "")
            woo = str(woo_raw).strip().lower()
            if not woo:
                continue
            if woo in name_mapping:
                continue
            zoho = "" if pd.isna(row.get("zoho name")) else str(row.get("zoho name")).strip()
            hsn_val = "" if pd.isna(row.get("hsn")) else str(row.get("hsn")).strip()
            usage_val = "" if pd.isna(row.get("usage unit")) else str(row.get("usage unit")).strip()
            name_mapping[woo] = {"zoho": zoho, "hsn": hsn_val, "usage_unit": usage_val}
        return name_mapping, item_db_df
    except FileNotFoundError:
        st.info("No item_database.xlsx found in app folder and no upload provided. Proceeding without mapping.")
        return {}, None
    except Exception as e:
        st.error(f"Error reading item database: {e}")
        try:
            ActivityLogger.log(
                user_id=SessionManager.get_user()['id'],
                action_type='module_error',
                module_key='woocommerce_zoho_export',
                description=f"Error reading item_database.xlsx: {e}",
                success=False
            )
        except Exception:
            pass
        return {}, None

def fetch_orders(api_url, consumer_key, consumer_secret, start_iso, end_iso):
    """
    Fetch orders with pagination and retry logic.
    Returns list of orders.
    """
    all_orders = []
    page = 1
    max_retries = 3
    retry_delay = 2

    while True:
        retries = 0
        while retries < max_retries:
            try:
                resp = requests.get(
                    f"{api_url.rstrip('/')}/orders",
                    params={
                        "after": start_iso,
                        "before": end_iso,
                        "per_page": 100,
                        "page": page,
                        "status": "any",
                        "order": "asc",
                        "orderby": "id"
                    },
                    auth=(consumer_key, consumer_secret),
                    timeout=30
                )

                # Rate limiting
                if resp.status_code == 429:
                    retry_after = int(resp.headers.get('Retry-After', 60))
                    st.warning(f"Rate limit reached. Waiting {retry_after} seconds before retrying...")
                    time.sleep(retry_after)
                    retries += 1
                    continue

                if resp.status_code != 200:
                    st.error(f"Error fetching orders: {resp.status_code} - {resp.text}")
                    return []

                orders = resp.json()
                if not isinstance(orders, list):
                    st.error("Invalid response format from WooCommerce API (expected list).")
                    return []

                if not orders:
                    return all_orders

                all_orders.extend(orders)
                page += 1
                break  # success -> exit retry loop

            except requests.exceptions.Timeout:
                st.error("Network timeout. Please check your connection and try again.")
                return []
            except requests.exceptions.ConnectionError:
                st.error("Network issue - Unable to connect to WooCommerce. Please try again.")
                return []
            except requests.exceptions.RequestException as e:
                if retries < max_retries - 1:
                    st.warning(f"Request failed. Retrying in {retry_delay} seconds... (Attempt {retries + 1}/{max_retries})")
                    time.sleep(retry_delay)
                    retries += 1
                else:
                    st.error(f"Network issue - {str(e)}. Please try again.")
                    return []
            except Exception as e:
                st.error(f"Unexpected error: {str(e)}. Please try again.")
                return []

    return all_orders

def transform_orders_to_rows(all_orders, name_mapping, invoice_prefix, start_sequence):
    """
    Transform completed orders into CSV rows and build replacements log.
    """
    all_orders.sort(key=lambda x: x.get("id", 0))
    completed_orders = [o for o in all_orders if o.get("status", "").lower() == "completed"]
    if not completed_orders:
        return [], [], []

    csv_rows = []
    replacements_log = []
    sequence_number = start_sequence

    for order in completed_orders:
        order_id = order.get("id")
        invoice_number = f"{invoice_prefix}{sequence_number:05d}"
        sequence_number += 1
        invoice_date = parse(order.get("date_created")).strftime("%Y-%m-%d %H:%M:%S") if order.get("date_created") else ""
        customer_name = f"{order.get('billing',{}).get('first_name','')} {order.get('billing',{}).get('last_name','')}".strip()
        place_of_supply = order.get('billing',{}).get('state','')
        currency = order.get('currency','')
        shipping_charge = to_float(order.get('shipping_total',0))
        entity_discount = to_float(order.get('discount_total',0))

        for item in order.get("line_items", []):
            product_meta = item.get("meta_data", []) or []
            # Try to extract HSN & usage unit from meta_data first
            hsn = ""
            usage_unit = ""
            for meta in product_meta:
                key = str(meta.get("key","")).lower()
                if key == "hsn":
                    hsn_val = meta.get("value","")
                    hsn = "" if hsn_val is None else str(hsn_val)
                if key == "usage unit":
                    usage_val = meta.get("value","")
                    usage_unit = "" if usage_val is None else str(usage_val)

            # Mapping using item_database
            original_item_name = item.get("name","")
            item_name_lower = str(original_item_name).strip().lower()
            if item_name_lower in name_mapping:
                mapping = name_mapping[item_name_lower]
                item_name_final = mapping.get("zoho", original_item_name) or original_item_name
                hsn_from_db = mapping.get("hsn", "")
                usage_from_db = mapping.get("usage_unit", "")
                if hsn_from_db:
                    hsn = hsn_from_db
                if usage_from_db:
                    usage_unit = usage_from_db
                replacements_log.append({
                    "Original WooCommerce Name": original_item_name,
                    "Replaced Zoho Name": item_name_final,
                    "HSN": hsn,
                    "Usage unit": usage_unit
                })
            else:
                item_name_final = original_item_name

            # Ensure numeric tax percent
            tax_class = item.get("tax_class") or ""
            try:
                item_tax_pct = float(tax_class)
            except (TypeError, ValueError):
                item_tax_pct = 0.0

            row = {
                "Invoice Number": invoice_number,
                "PurchaseOrder": order_id,
                "Invoice Date": invoice_date,
                "Invoice Status": order.get("status", "").capitalize(),
                "Customer Name": customer_name,
                "Place of Supply": place_of_supply,
                "Currency Code": currency,
                "Item Name": item_name_final,
                "HSN/SAC": hsn,
                "Item Type": item.get("type", "goods"),
                "Quantity": item.get("quantity", 0),
                "Usage unit": usage_unit,
                "Item Price": to_float(item.get("price", 0)),
                "Is Inclusive Tax": "FALSE",
                "Item Tax %": item_tax_pct,
                "Discount Type": "entity_level",
                "Is Discount Before Tax": "TRUE",
                "Entity Discount Amount": entity_discount,
                "Shipping Charge": shipping_charge,
                "Item Tax Exemption Reason": "ITEM EXEMPT FROM GST",
                "Supply Type": "Exempted",
                "GST Treatment": "consumer"
            }
            csv_rows.append(row)

    return csv_rows, replacements_log, completed_orders

def build_summary_and_order_details(completed_orders, invoice_prefix, start_sequence):
    """
    Build summary metrics DataFrame and order details DataFrame.
    """
    all_orders_count = len(completed_orders)
    first_order_id = completed_orders[0].get("id") if completed_orders else None
    last_order_id = completed_orders[-1].get("id") if completed_orders else None
    first_invoice_number = f"{invoice_prefix}{start_sequence:05d}"
    last_invoice_number = f"{invoice_prefix}{(start_sequence + len(completed_orders) - 1):05d}" if completed_orders else None

    total_revenue_by_order_total = 0.0
    order_details_rows = []
    seq_temp = start_sequence
    for order in completed_orders:
        order_total = to_float(order.get("total", 0))
        refunds = order.get("refunds") or []
        refund_total = sum(to_float(r.get("amount") or r.get("total") or r.get("refund_total") or 0) for r in refunds)
        net_total = order_total - refund_total
        total_revenue_by_order_total += net_total

        invoice_number_temp = f"{invoice_prefix}{seq_temp:05d}"
        seq_temp += 1
        order_details_rows.append({
            "Invoice Number": invoice_number_temp,
            "Order Number": order.get("id"),
            "Date": parse(order.get("date_created")).strftime("%Y-%m-%d %H:%M:%S") if order.get("date_created") else "",
            "Customer Name": f"{order.get('billing',{}).get('first_name','')} {order.get('billing',{}).get('last_name','')}".strip(),
            "Order Total": net_total
        })

    summary_metrics = {
        "Metric": [
            "Total Orders Fetched",
            "Completed Orders",
            "Total Revenue (Net of Refunds)",
            "Completed Order ID Range",
            "Invoice Number Range"
        ],
        "Value": [
            all_orders_count,
            all_orders_count,
            total_revenue_by_order_total,
            f"{first_order_id} → {last_order_id}" if completed_orders else "",
            f"{first_invoice_number} → {last_invoice_number}" if completed_orders else ""
        ]
    }
    summary_df = pd.DataFrame(summary_metrics)
    order_details_df = pd.DataFrame(order_details_rows)
    grand_total = order_details_df["Order Total"].sum() if not order_details_df.empty else 0.0
    grand_total_row = {
        "Invoice Number": "Grand Total",
        "Order Number": "",
        "Date": "",
        "Customer Name": "",
        "Order Total": grand_total
    }
    if not order_details_df.empty:
        order_details_df = pd.concat([order_details_df, pd.DataFrame([grand_total_row])], ignore_index=True)
    else:
        order_details_df = pd.DataFrame([grand_total_row])

    return summary_df, order_details_df

def create_excel_bytes(summary_df, order_details_df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary Metrics")
        order_details_df.to_excel(writer, index=False, sheet_name="Order Details")
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            # Header formatting
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            # Column widths
            for col in ws.columns:
                max_length = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length
    return output.getvalue()

def create_zip_bytes(csv_bytes, excel_bytes, start_date, end_date):
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, "w") as zip_file:
        zip_file.writestr(f"orders_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv", csv_bytes)
        zip_file.writestr(f"summary_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx", excel_bytes)
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

# ------------------------
# Main module show()
def show():
    """
    Streamlit entry point for the woocommerce_zoho_export module.
    """
    # Module access check
    SessionManager.require_module_access('woocommerce_zoho_export')

    user = SessionManager.get_user()
    profile = SessionManager.get_user_profile()

    st.markdown("### 📦 WooCommerce → Zoho Export")
    st.markdown("Fetch completed orders from WooCommerce and export CSV + Excel bundled into a ZIP (`orders_export.zip`).")
    st.markdown("---")

    # Read credentials from st.secrets["woocommerce"]
    try:
        WC_API_URL = st.secrets["woocommerce"]["api_url"]
        WC_CONSUMER_KEY = st.secrets["woocommerce"]["consumer_key"]
        WC_CONSUMER_SECRET = st.secrets["woocommerce"]["consumer_secret"]
    except Exception:
        st.error("⚠️ WooCommerce API credentials are missing from secrets! Expected keys under [woocommerce]: api_url, consumer_key, consumer_secret")
        st.info("""
        **Required secrets:**
        ```toml
        [woocommerce]
        api_url = "https://your-site.com/wp-json/wc/v3"
        consumer_key = "ck_xxxxx"
        consumer_secret = "cs_xxxxx"
        ```
        """)
        try:
            ActivityLogger.log(
                user_id=user['id'],
                action_type='module_error',
                module_key='woocommerce_zoho_export',
                description="Missing WooCommerce credentials in st.secrets",
                success=False
            )
        except Exception:
            pass
        return

    # Item DB upload (optional)
    st.markdown("#### Item Database (optional)")
    st.markdown("Upload `item_database.xlsx` with columns: WooCommerce Name, Zoho Name, HSN, Usage Unit (case-insensitive). If not provided, the app will try to read from app folder.")
    uploaded_item_db = st.file_uploader("Upload item_database.xlsx (optional)", type=['xlsx', 'xls'])
    name_mapping, item_db_df = read_item_database(uploaded_item_db)
    if item_db_df is not None:
        with st.expander("Preview Item Database"):
            st.dataframe(item_db_df, use_container_width=True)

    # Date inputs
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date")
    with col2:
        end_date = st.date_input("End Date")

    if start_date > end_date:
        st.error("Start date cannot be after end date.")
        return

    # Invoice prefix and starting sequence (no default for sequence)
    invoice_prefix = st.text_input("Invoice Prefix", value="ECHE/2526/")
    start_sequence_input = st.text_input("Starting Sequence Number (required - no default)", value="")

    # Fetch button and logs container
    fetch_button = st.button("Fetch Orders", disabled=False)
    log_container = st.empty()
    logs = []

    def append_log(msg, lvl="info"):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{timestamp}] {lvl.upper()}: {msg}"
        logs.append(line)
        log_container.text_area("Logs", value="\n".join(logs), height=240)

    if fetch_button:
        # Validate starting sequence
        if not start_sequence_input or start_sequence_input.strip() == "":
            st.error("Please enter a Starting Sequence Number (required).")
            append_log("Missing starting sequence number.", "error")
            try:
                ActivityLogger.log(
                    user_id=user['id'],
                    action_type='module_error',
                    module_key='woocommerce_zoho_export',
                    description="Missing starting sequence number",
                    success=False
                )
            except Exception:
                pass
            return
        try:
            start_sequence = int(start_sequence_input)
            if start_sequence < 1:
                raise ValueError("Sequence must be >= 1")
        except Exception as e:
            st.error(f"Invalid starting sequence number: {e}")
            append_log(f"Invalid starting sequence number: {e}", "error")
            try:
                ActivityLogger.log(
                    user_id=user['id'],
                    action_type='module_error',
                    module_key='woocommerce_zoho_export',
                    description=f"Invalid starting sequence number: {e}",
                    success=False
                )
            except Exception:
                pass
            return

        append_log("Starting WooCommerce export...", "info")
        try:
            ActivityLogger.log(
                user_id=user['id'],
                action_type='module_use',
                module_key='woocommerce_zoho_export',
                description=f"User started WooCommerce export for {start_date} to {end_date}"
            )
        except Exception:
            pass

        start_iso = start_date.strftime("%Y-%m-%dT00:00:00")
        end_iso = end_date.strftime("%Y-%m-%dT23:59:59")

        # Fetch
        try:
            all_orders = fetch_orders(WC_API_URL, WC_CONSUMER_KEY, WC_CONSUMER_SECRET, start_iso, end_iso)
            append_log(f"Fetched {len(all_orders)} orders from WooCommerce.", "info")
        except Exception as e:
            st.error(f"Error fetching orders: {e}")
            append_log(f"Error fetching orders: {e}", "error")
            try:
                ActivityLogger.log(
                    user_id=user['id'],
                    action_type='module_error',
                    module_key='woocommerce_zoho_export',
                    description=f"Error fetching orders: {e}",
                    success=False
                )
            except Exception:
                pass
            return

        if not all_orders:
            st.warning("No orders found in this date range.")
            append_log("No orders found in this date range.", "info")
            try:
                ActivityLogger.log(
                    user_id=user['id'],
                    action_type='module_use',
                    module_key='woocommerce_zoho_export',
                    description=f"No orders found for {start_date} to {end_date}"
                )
            except Exception:
                pass
            return

        # Transform
        csv_rows, replacements_log, completed_orders = transform_orders_to_rows(
            all_orders, name_mapping, invoice_prefix, start_sequence
        )

        if not completed_orders:
            st.warning("No completed orders found in this date range.")
            append_log("No completed orders found.", "info")
            try:
                ActivityLogger.log(
                    user_id=user['id'],
                    action_type='module_use',
                    module_key='woocommerce_zoho_export',
                    description=f"No completed orders for {start_date} to {end_date}"
                )
            except Exception:
                pass
            return

        df = pd.DataFrame(csv_rows)
        st.subheader("Line Items Preview (first 50 rows)")
        st.dataframe(df.head(50), use_container_width=True)

        if replacements_log:
            st.subheader("Item Name Replacements Log")
            st.dataframe(pd.DataFrame(replacements_log), use_container_width=True)
            append_log(f"Applied {len(replacements_log)} item name replacements.", "info")

        # Summary & Order details
        summary_df, order_details_df = build_summary_and_order_details(completed_orders, invoice_prefix, start_sequence)
        st.subheader("Summary Metrics")
        st.dataframe(summary_df, use_container_width=True)

        # Prepare files
        try:
            csv_bytes = df.to_csv(index=False).encode('utf-8')
            excel_bytes = create_excel_bytes(summary_df, order_details_df)
            zip_bytes = create_zip_bytes(csv_bytes, excel_bytes, start_date, end_date)
        except Exception as e:
            st.error(f"Error preparing exports: {e}")
            append_log(f"Error preparing exports: {e}", "error")
            try:
                ActivityLogger.log(
                    user_id=user['id'],
                    action_type='module_error',
                    module_key='woocommerce_zoho_export',
                    description=f"Error preparing exports: {e}",
                    success=False
                )
            except Exception:
                pass
            return

        # Success
        append_log(f"Export prepared successfully. Orders exported: {len(completed_orders)}", "info")
        st.success(f"Export ready — {len(completed_orders)} completed orders. Download below.")
        try:
            ActivityLogger.log(
                user_id=user['id'],
                action_type='module_use',
                module_key='woocommerce_zoho_export',
                description=f"Exported {len(completed_orders)} completed orders from {start_date} to {end_date}",
                metadata={'orders_exported': len(completed_orders), 'date_from': str(start_date), 'date_to': str(end_date)}
            )
        except Exception:
            pass

        # Provide ZIP download (fixed name orders_export.zip)
        st.download_button(
            label="Download CSV + Excel (Combined ZIP)",
            data=zip_bytes,
            file_name="orders_export.zip",
            mime="application/zip"
        )
